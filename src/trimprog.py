# coding: utf-8
# Roma e20
# Written by Luca Allulli

# Requires pupeteer-pdf. Install it globally:
# sudo npm i -g puppeteer-pdf

# Also requires pdftk
import shutil
from datetime import datetime, timedelta
from contextlib import contextmanager
from openpyxl import load_workbook
from utils import skipping_iter, chdir, create_dir_if_not_existing
import subprocess
import os, sys
import jinja2
# import pdfkit

DAYS = ['Lu', 'Ma', 'Me', 'Gio', 'Ve', 'Sa', 'Do']
PROGRAM_WS = 'Programma'
TYPES_WS = 'Classi'
NOTES = 'Note'
NUMBER_OF_COPIES = 100


class TrimProg(object):
	def __init__(self):
		# Map type code to dict with keys 'desc', 'class'
		self.types = {}
		# Name of units, in order
		self.units = []
		# Program: list of items with following keys:
		# - date: date or special text
		# - formatted_date: formatted date or special text
		# - is_date: boolean
		# - is_festive: boolean
		# - eom: boolean, end of month (last date of a month)
		# - programs: list of lists of program codes -- one list per unit
		# - notes: notes
		self.program = []
		self.year = None

	def _load_types_from_wb(self, wb):
		ws = wb[TYPES_WS]
		for row in skipping_iter(ws):
			code, desc, klass = row[:3]
			self.types[code.value] = {
				'desc': desc.value,
				'class': klass.value
			}

	def _load_units_from_ws(self, rows):
		notes = NOTES.lower()
		for c in skipping_iter(rows[0], 3):
			v = c.value
			if v.lower() == notes:
				break
			self.units.append(v)

	def _get_type(self, code):
		if code in self.types:
			return self.types[code]
		return {
			'desc': code,
			'class': '',
		}

	def _get_value(self, c):
		v = c.value
		if v is None:
			return ''
		return v

	def _load_program_from_ws(self, rows):
		n = len(self.units)
		last_date = None
		last_month = None
		for row in skipping_iter(rows):
			row = row[:n + 4]
			vrow = [self._get_value(r) for r in row] + [""] * (n + 4 - len(row))
			date = vrow[0]
			# Skip blank rows or rows beginning with '#'
			if date == '' or isinstance(date, str) and date.strip()[0] == '#':
				continue
			fest = vrow[1]
			eom = True if vrow[2] == 1 else False
			acts = vrow[3:n + 3]
			notes = vrow[n + 3]
			if last_date != date:
				is_festive = False
				is_date = row[0].is_date
				if fest == 1 or (is_date and date.weekday() == 6):
					is_festive = True
				if is_date:
					wd = date.weekday()
					if False: # wd == 6:
						yesterday = date - timedelta(days=1)
						formatted_date = u"Sa {} e Do {}".format(yesterday.day, date.strftime('%d/%m'))
					else:
						formatted_date = u"{} {}".format(DAYS[wd], date.strftime('%d/%m'))
					if last_month is not None and last_month != date.month:
						eom = True
					last_month = date.month
					if self.year is None:
						self.year = date.year
				else:
					formatted_date = date
				self.program.append({
					'date': date,
					'formatted_date': formatted_date,
					'is_festive': is_festive,
					'is_date': is_date,
					'programs': [[self._get_type(a)] for a in acts],
					'notes': notes,
					'eom': eom,
				})
				last_date = date
			else:
				p = self.program[-1]['programs']
				for i in range(len(acts)):
					p[i].append(self._get_type(acts[i]))

	def get_first_date(self):
		for p in self.program:
			if p['is_date']:
				return p['date']

	def get_quarter(self):
		d = self.get_first_date()
		y = d.year
		m = d.month
		if m <= 3:
			q = 'Q1'
		elif m <= 8:
			q = 'Q2'
		else:
			q = 'Q4'
		return f"{y}-{q}"

	def load_workbook(self, filename):
		wb = load_workbook(filename, read_only=True)
		self._load_types_from_wb(wb)
		rows = list(wb[PROGRAM_WS])
		self._load_units_from_ws(rows)
		self._load_program_from_ws(rows)
		wb.close()


def jinja2_render(tpl_path, context):
	path, filename = os.path.split(tpl_path)
	return jinja2.Environment(
		loader=jinja2.FileSystemLoader(path or './')
	).get_template(filename).render(context)


def render_web(tp):
	"""
	Render a TrimProg as a web page

	:param tp: TrimProg instance
	:return: HTML code
	"""
	ctx = {
		'units': [x.replace('\n', '<br />') for x in tp.units],
		'program': tp.program,
		'year': tp.year,
		'today': datetime.today(),
	}
	return jinja2_render('templates/roma_e20.html', ctx)


def pdf_filename(filename):
	fn, ext = os.path.splitext(filename)
	return fn + ".pdf"


def print_to_pdf(filename):
	print("\nPrinting using pupeteer-pdf")
	pdf_name = pdf_filename(filename)
	subprocess.call([
		"puppeteer-pdf",
		filename,
		"--path", pdf_name,
		"-s", "0.8",
		"-f", "A4",
		"--printBackground",
	])


def create_copies(filename, number=NUMBER_OF_COPIES):
	fn, ext = os.path.splitext(filename)
	pdf_name = fn + ".pdf"
	fn_n = fn + f" x{number}"
	pdf_n = fn_n + ".pdf"
	copies = ["A"] * number
	print("Duplicating file")
	args = [
		"pdftk",
		f'A={pdf_name}',
		"cat",
	] + copies + [
		"output", pdf_n,
	]
	subprocess.call(args)
	print("Compressing to 7zip")
	subprocess.call([
		"7z",
		"a",
		f"{fn_n}.7z",
		pdf_n,
	])
	os.remove(pdf_n)


# def printToPdf(filename):
# 	options = {
# 		'page-size': 'A4',
# 		'enable-local-file-access': True,
# 		'dpi': 50,
# 	}
# 	pdfkit.from_file('Roma e20.html', 'Roma e20.pdf', options=options)


def main(pdf=False):
	tp = TrimProg()
	tp.load_workbook('Template.xlsx')
	quarter = tp.get_quarter()
	print(f"Quarter: {quarter}")
	create_dir_if_not_existing(quarter)
	filename = f'Roma e20-{quarter}.html'
	html = render_web(tp)
	with open(filename, 'w') as o:
		o.write(html)
	if pdf:
		print_to_pdf(filename)
		pdf = pdf_filename(filename)
		os.replace(pdf, os.path.join(quarter, pdf))
	os.replace(filename, os.path.join(quarter, filename))
	shutil.copy("Template.xlsx", quarter)
	os.replace(
		os.path.join(quarter, "Template.xlsx"),
		os.path.join(quarter, f"Template-{quarter}.xlsx"),
	)

	with chdir(quarter):
		print("If you print PDF from Brave:")
		print("1. Scale: 80%")
		print("2. Don't place mouse pointer over the table (use CTRL-P to open print dialog)")
		create_copies(filename)


if __name__ == '__main__':
	# main(True)
	main('pdf' in sys.argv)

